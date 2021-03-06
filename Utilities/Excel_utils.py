
'''
import openpyxl


def getRowCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column

def getCellData(path,sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value

def setCellData(path,sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value= data
    workbook.save(path) # To save workbook / sheet


userPath = "../Excel/Test_Data_Python.xlsx"
userSheetName = "Sheet1"

rows = getRowCount(userPath,userSheetName)
cols = getColCount(userPath,userSheetName)

print("Total rows: ", rows)
print("Total columns: ", cols)

print(getCellData(userPath,userSheetName,1,2))

setCellData(userPath,userSheetName,1,4, "Hello")

'''

import openpyxl

def read_data_from_excel(file_name, sheet):
    workbook = openpyxl.load_workbook(filename=file_name)
    sheet = workbook[sheet]

    max_rows = sheet.max_row
    max_clm = sheet.max_column

    print("Max row number:", max_rows)
    print("Max clm number:", max_clm)

    for r in range(1, max_rows + 1):
        for c in range(1, max_clm+1):
            print(sheet.cell(row=r, column=c).value, end="  ")
        print()
